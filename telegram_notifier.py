# -*- coding: utf-8 -*-
"""
Telegram Notifier for ZKTeco Attendance Dashboard
==================================================
Features:
  1. Device online / offline alerts
  2. Per-punch notifications (IP, time, employee code)
  3. Daily 08:10 absent report (XLSX document, grouped by dept category)
  4. Bot command handler:
       device status, device sync, device reboot, user search,
       today summary, today absent, dept summary, cache refresh,
       user report, unknown users, pending punches, holiday check,
       db stats

Configuration (settings.ini [telegram] section):
  bot_token   = <your bot token>
  chat_id     = <target chat / group id>
  enabled     = 1
  notify_device_status = 1
  notify_punches       = 1
  notify_daily_report  = 1
  daily_report_hour    = 8
  daily_report_minute  = 10
"""

import io
import logging
import threading
import time
from datetime import datetime
from typing import Optional, Dict, Callable, Any

import httpx

logger = logging.getLogger("ZKTeco.Telegram")


# ---------------------------------------------------------------------------
#  Low-level HTTP helpers (fully synchronous, no asyncio dependency)
# ---------------------------------------------------------------------------

def _post(url: str, **kwargs) -> Optional[dict]:
    """POST to Telegram API; return parsed JSON or None on network error.

    Non-200 responses are also returned as dicts (containing ``ok=false`` and
    a ``description`` field) so callers can surface the exact Telegram error.
    """
    try:
        with httpx.Client(timeout=15) as client:
            resp = client.post(url, **kwargs)
        try:
            data = resp.json()
        except ValueError:
            data = {}
        if resp.status_code != 200:
            logger.warning("Telegram API %s: status %s  %s", url, resp.status_code, resp.text[:200])
        return data
    except Exception as exc:
        logger.warning("Telegram send error: %s", exc)
    return None


def _format_punch_time(ts: str) -> str:
    """Convert 'YYYY-MM-DD HH:MM:SS' to 'DD-Mon-YYYY HH:MM:SSAM/PM'."""
    try:
        dt = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%d-%b-%Y %I:%M:%S%p")
    except Exception:
        return ts


# ---------------------------------------------------------------------------
#  TelegramNotifier
# ---------------------------------------------------------------------------

class TelegramNotifier:
    """
    Telegram bot notifier for the ZKTeco attendance dashboard.
    All send_* methods are synchronous and thread-safe.
    """

    def __init__(
        self,
        bot_token: str,
        chat_id: str,
        enabled: bool = True,
        notify_device_status: bool = True,
        notify_punches: bool = True,
        notify_daily_report: bool = True,
        system_name: str = "Attendance",
        notification_settings: dict = None,
    ):
        # notification_settings can override individual notify_* flags when provided
        if notification_settings:
            notify_device_status = notification_settings.get("device_status", notify_device_status)
            notify_punches       = notification_settings.get("punches",       notify_punches)
            notify_daily_report  = notification_settings.get("daily_report",  notify_daily_report)
        self.bot_token = bot_token.strip() if bot_token else ""
        self.chat_id = str(chat_id).strip() if chat_id else ""
        self.enabled = enabled
        self.notify_device_status = notify_device_status
        self.notify_punches = notify_punches
        self.notify_daily_report = notify_daily_report
        self.system_name = system_name
        self._base = "https://api.telegram.org/bot{0}".format(self.bot_token)
        self._lock = threading.Lock()

    # ------------------------------------------------------------------ #
    #  Internal plumbing                                                   #
    # ------------------------------------------------------------------ #

    def _ok(self) -> bool:
        return bool(self.enabled and self.bot_token and self.chat_id)

    def _send_message(self, text: str, parse_mode: str = "HTML") -> bool:
        """Send a plain text message."""
        if not self._ok():
            return False
        # Telegram HTML has a 4096-char limit; truncate gracefully
        if len(text) > 4000:
            text = text[:3990] + "\n…"
        with self._lock:
            result = _post(
                self._base + "/sendMessage",
                json={"chat_id": self.chat_id, "text": text, "parse_mode": parse_mode},
            )
        return result is not None and result.get("ok", False)

    def _send_document(self, file_bytes: bytes, filename: str, caption: str = "") -> bool:
        """Upload a binary file as a Telegram document."""
        if not self._ok():
            return False
        with self._lock:
            result = _post(
                self._base + "/sendDocument",
                data={"chat_id": self.chat_id, "caption": caption[:1024]},
                files={"document": (filename, file_bytes,
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        return result is not None and result.get("ok", False)

    # ------------------------------------------------------------------ #
    #  Feature 1: Device online / offline                                  #
    # ------------------------------------------------------------------ #

    def notify_device_online(self, ip: str, name: str = "") -> bool:
        if not (self._ok() and self.notify_device_status):
            return False
        label = "{0} ({1})".format(name, ip) if name else ip
        msg = (
            "✅ <b>{sys}</b> — Device ONLINE\n"
            "📡 <b>{label}</b>\n"
            "🕐 {ts}"
        ).format(sys=self.system_name, label=label,
                 ts=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        return self._send_message(msg)

    def notify_device_offline(self, ip: str, name: str = "", error: str = "") -> bool:
        if not (self._ok() and self.notify_device_status):
            return False
        label = "{0} ({1})".format(name, ip) if name else ip
        err_line = "\n⚠️ {0}".format(error) if error else ""
        msg = (
            "🔴 <b>{sys}</b> — Device OFFLINE\n"
            "📡 <b>{label}</b>\n"
            "🕐 {ts}{err}"
        ).format(sys=self.system_name, label=label,
                 ts=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                 err=err_line)
        return self._send_message(msg)

    # ------------------------------------------------------------------ #
    #  Feature 2: Per-punch notification                                   #
    # ------------------------------------------------------------------ #

    def notify_punch(self, badge: str, name: str, ip: str, ts: str) -> bool:
        if not (self._ok() and self.notify_punches):
            return False
        msg = (
            "🪪 <b>{badge}</b> : {name}\n"
            "🕐 {ts}\n"
            "📡 Device: <code>{ip}</code>"
        ).format(badge=badge, name=name or "Unknown", ip=ip,
                 ts=_format_punch_time(ts))
        return self._send_message(msg)

    # ------------------------------------------------------------------ #
    #  Feature 3 & 4: Daily absent report at 08:10                        #
    # ------------------------------------------------------------------ #

    def send_daily_absent_report(
        self,
        absent: list,
        present_count: int,
        total: int,
        date_str: str,
        dept_order: list,
    ) -> bool:
        """
        Send the daily absent report.
        Sends a short summary text message AND an XLSX attachment
        grouped by dept category:
          Teachers | Admin+Support | Drivers+Conductors | Cleaners | Others
        """
        if not (self._ok() and self.notify_daily_report):
            return False

        absent_count = len(absent)

        # ---------- 1. Text summary ----------
        # Group absent by dept for the inline summary
        dept_buckets: Dict[str, list] = {}
        for emp in absent:
            dept_buckets.setdefault(emp.get("dept", "OTHER"), []).append(emp)

        lines = [
            "📋 <b>{sys} — Daily Absent Report</b>".format(sys=self.system_name),
            "📅 {date}".format(date=date_str),
            "❌ Absent: <b>{a}</b>  ✅ Present: <b>{p}</b>  👥 Total: <b>{t}</b>".format(
                a=absent_count, p=present_count, t=total),
            "",
        ]

        # Use configured dept order for display priority
        ordered_depts = [d for d in (dept_order or []) if d in dept_buckets]
        ordered_depts += sorted(k for k in dept_buckets if k not in ordered_depts)
        for dept in ordered_depts:
            emps = dept_buckets[dept]
            lines.append("<b>{dept}</b> ({n} absent)".format(dept=dept, n=len(emps)))
            for emp in sorted(emps, key=lambda e: e.get("name", "")):
                lines.append("  · {code}  {name}".format(
                    code=emp.get("code", ""), name=emp.get("name", "")))
            lines.append("")

        msg_ok = self._send_message("\n".join(lines))

        # ---------- 2. XLSX attachment ----------
        try:
            xlsx_bytes = _build_absent_xlsx(absent, date_str)
            fname = "absent_{0}.xlsx".format(
                datetime.now().strftime("%Y%m%d"))
            caption = "Absent report {0} — {1} absent / {2} total".format(
                date_str, absent_count, total)
            self._send_document(xlsx_bytes, fname, caption)
        except Exception as exc:
            logger.warning("Could not build XLSX for Telegram: %s", exc)

        return msg_ok

    # ------------------------------------------------------------------ #
    #  Utility                                                             #
    # ------------------------------------------------------------------ #

    def test_connection(self) -> tuple:
        """Send a test message to verify the bot is working.

        Returns ``(ok: bool, message: str)`` where *message* is either a
        success note or the exact error description returned by the Telegram
        API (e.g. "Unauthorized", "Bad Request: chat not found").
        """
        if not self._ok():
            if not self.bot_token:
                return False, "Bot token not set"
            if not self.chat_id:
                return False, "Chat ID not set"
            return False, "Telegram notifications are disabled"

        # Step 1 — validate the token before attempting to send
        me = _post(self._base + "/getMe")
        if me is None:
            return False, "Could not reach Telegram API — check network connectivity"
        if not me.get("ok"):
            desc = me.get("description", "Unknown error")
            return False, "Invalid bot token: {0}".format(desc)

        # Step 2 — try to send the test message
        msg = (
            "🧪 <b>{sys} — Test</b>\n"
            "📅 {ts}\n"
            "✅ Telegram bot is connected and working."
        ).format(sys=self.system_name, ts=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        with self._lock:
            result = _post(
                self._base + "/sendMessage",
                json={"chat_id": self.chat_id, "text": msg, "parse_mode": "HTML"},
            )
        if result is None:
            return False, "Could not reach Telegram API — check network connectivity"
        if not result.get("ok"):
            return False, result.get("description", "Send failed")
        return True, "Test message sent successfully"

    # Kept for backward compatibility with old code
    def send_message_sync(self, message: str, parse_mode: str = "HTML") -> bool:
        return self._send_message(message, parse_mode)


# ---------------------------------------------------------------------------
#  TelegramBotHandler — interactive command handler (long-polling)
# ---------------------------------------------------------------------------

class TelegramBotHandler:
    """
    Polls Telegram for incoming messages in a background thread and handles
    bot commands typed in the configured chat:

      device status   — show all device online/offline, punches today, user count
      device sync     — sync time & users across all devices
      device reboot   — present inline keyboard to pick a device (or all)
      user search     — prompt for employee name/badge and show punch timings today
      user <name/badge>  — directly look up employee by name or badge (e.g. "user 1024", "user john")
      today summary   — present/absent/total counts and cache age
      today absent    — full absent list grouped by department
      dept summary    — per-department present vs absent breakdown
      cache refresh   — trigger an immediate attendance-data refresh
      user report     — punch times for a specific employee today
      unknown users   — badge IDs seen on devices but not in employee list
      pending punches — punch-correction requests awaiting admin approval
      holiday check   — today's holiday status + upcoming holidays (30 days)
      db stats        — employee count, punch records, DB size
    """

    def __init__(
        self,
        bot_token: str,
        chat_id: str,
        get_today_fn: Callable = None,
        get_device_ips_fn: Callable = None,
        get_device_names_fn: Callable = None,
        get_device_status_fn: Callable = None,
        sync_clocks_fn: Callable = None,
        sync_users_fn: Callable = None,
        reboot_device_fn: Callable = None,
        search_employee_fn: Callable = None,
        get_db_stats_fn: Callable = None,
        get_unknown_users_fn: Callable = None,
        get_upcoming_holidays_fn: Callable = None,
        get_pending_punches_fn: Callable = None,
        cache_refresh_fn: Callable = None,
        get_employee_punches_fn: Callable = None,
        find_employee_fn: Callable = None,
        get_punch_records_fn: Callable = None,
        get_late_today_fn: Callable = None,
        get_early_exits_fn: Callable = None,
        get_week_summary_fn: Callable = None,
        get_month_dept_summary_fn: Callable = None,
        get_top_absent_fn: Callable = None,
        get_who_is_in_fn: Callable = None,
        get_punch_feed_fn: Callable = None,
    ):
        self.bot_token = bot_token.strip() if bot_token else ""
        self.chat_id = str(chat_id).strip() if chat_id else ""
        self._base = "https://api.telegram.org/bot{0}".format(self.bot_token)
        self._offset = 0
        self._running = False
        self._thread = None
        # Conversation state per chat_id: {"state": str}
        self._state: Dict[str, dict] = {}

        # Callbacks supplied by the host (server.py)
        self.get_today_fn = get_today_fn
        self.get_device_ips_fn = get_device_ips_fn
        self.get_device_names_fn = get_device_names_fn
        self.get_device_status_fn = get_device_status_fn
        self.sync_clocks_fn = sync_clocks_fn
        self.sync_users_fn = sync_users_fn
        self.reboot_device_fn = reboot_device_fn
        self.search_employee_fn = search_employee_fn
        self.get_db_stats_fn = get_db_stats_fn
        self.get_unknown_users_fn = get_unknown_users_fn
        self.get_upcoming_holidays_fn = get_upcoming_holidays_fn
        self.get_pending_punches_fn = get_pending_punches_fn
        self.cache_refresh_fn = cache_refresh_fn
        self.get_employee_punches_fn = get_employee_punches_fn
        self.find_employee_fn = find_employee_fn
        self.get_punch_records_fn = get_punch_records_fn
        self.get_late_today_fn = get_late_today_fn
        self.get_early_exits_fn = get_early_exits_fn
        self.get_week_summary_fn = get_week_summary_fn
        self.get_month_dept_summary_fn = get_month_dept_summary_fn
        self.get_top_absent_fn = get_top_absent_fn
        self.get_who_is_in_fn = get_who_is_in_fn
        self.get_punch_feed_fn = get_punch_feed_fn

    # Max rows to include in list-type bot replies (to stay within Telegram's 4096-char limit)
    _MAX_LIST_ITEMS = 30

    def start(self):
        if not self.bot_token or not self.chat_id:
            logger.warning("[BotHandler] Cannot start — missing bot_token or chat_id")
            return
        self._running = True
        self._thread = threading.Thread(target=self._poll_loop, name="TgBotPoll", daemon=True)
        self._thread.start()
        logger.info("[BotHandler] Long-poll thread started (chat %s)", self.chat_id)

    def stop(self):
        self._running = False

    # ------------------------------------------------------------------ #
    #  Low-level helpers                                                   #
    # ------------------------------------------------------------------ #

    def _get_updates(self) -> Optional[dict]:
        try:
            with httpx.Client(timeout=35) as client:
                resp = client.post(
                    self._base + "/getUpdates",
                    json={
                        "offset": self._offset,
                        "timeout": 30,
                        "allowed_updates": ["message", "callback_query"],
                    },
                )
                return resp.json()
        except Exception as exc:
            logger.warning("[BotHandler] getUpdates error: %s", exc)
            return None

    def _send(self, chat_id: str, text: str, reply_markup: dict = None):
        if len(text) > 4000:
            text = text[:3990] + "\n…"
        payload: Dict[str, Any] = {
            "chat_id": chat_id,
            "text": text,
            "parse_mode": "HTML",
        }
        if reply_markup:
            payload["reply_markup"] = reply_markup
        _post(self._base + "/sendMessage", json=payload)

    def _answer_callback(self, callback_query_id: str, text: str = ""):
        _post(
            self._base + "/answerCallbackQuery",
            json={"callback_query_id": callback_query_id, "text": text},
        )

    def _edit_message_text(self, chat_id: str, message_id: int, text: str):
        _post(
            self._base + "/editMessageText",
            json={"chat_id": chat_id, "message_id": message_id,
                  "text": text, "parse_mode": "HTML"},
        )

    def _edit_message_text_and_keyboard(self, chat_id: str, message_id: int,
                                        text: str, reply_markup: dict = None):
        payload: Dict[str, Any] = {
            "chat_id": chat_id, "message_id": message_id,
            "text": text, "parse_mode": "HTML",
        }
        if reply_markup:
            payload["reply_markup"] = reply_markup
        _post(self._base + "/editMessageText", json=payload)

    # ------------------------------------------------------------------ #
    #  Polling loop                                                        #
    # ------------------------------------------------------------------ #

    def _poll_loop(self):
        while self._running:
            try:
                data = self._get_updates()
                if data and data.get("ok"):
                    for update in data.get("result", []):
                        self._offset = update["update_id"] + 1
                        try:
                            self._handle_update(update)
                        except Exception as exc:
                            logger.warning("[BotHandler] Update handling error: %s", exc)
                elif data is not None and not data.get("ok"):
                    logger.warning("[BotHandler] getUpdates not ok: %s", data.get("description", ""))
                    time.sleep(10)
            except Exception as exc:
                logger.warning("[BotHandler] Poll loop error: %s", exc)
                time.sleep(5)

    # ------------------------------------------------------------------ #
    #  Update dispatcher                                                   #
    # ------------------------------------------------------------------ #

    def _handle_update(self, update: dict):
        # Callback queries (inline keyboard button presses)
        if "callback_query" in update:
            cq = update["callback_query"]
            chat_id = str(cq["message"]["chat"]["id"])
            if chat_id != self.chat_id:
                return
            self._handle_callback(cq)
            return

        if "message" not in update:
            return
        msg = update["message"]
        chat_id = str(msg["chat"]["id"])
        if chat_id != self.chat_id:
            return

        raw_text = (msg.get("text") or "").strip()
        text_lower = raw_text.lower()

        # Handle multi-step conversation states first
        state_info = self._state.get(chat_id, {})
        if state_info.get("state") == "awaiting_search_query":
            self._state.pop(chat_id, None)
            self._handle_search_query(chat_id, raw_text)
            return
        if state_info.get("state") == "awaiting_user_report_query":
            self._state.pop(chat_id, None)
            self._handle_user_report_query(chat_id, raw_text)
            return
        if state_info.get("state") == "awaiting_attendance_query":
            self._state.pop(chat_id, None)
            self._handle_attendance_query(chat_id, raw_text)
            return

        # Single-step commands
        if text_lower in ("device status", "/device_status"):
            self._cmd_device_status(chat_id)
        elif text_lower in ("device sync", "/device_sync"):
            self._cmd_device_sync(chat_id)
        elif text_lower in ("device reboot", "/device_reboot"):
            self._cmd_device_reboot_ask(chat_id)
        elif text_lower in ("user search", "/user_search"):
            self._cmd_device_search_ask(chat_id)
        elif text_lower in ("user report", "/user_report"):
            self._cmd_user_report_ask(chat_id)
        elif text_lower in ("user attendance", "user att", "/user_attendance"):
            self._cmd_user_attendance_ask(chat_id)
        elif text_lower.startswith("user ") and len(text_lower) > 5:
            # Direct lookup: "user 1024" or "user showkath"
            self._cmd_user_direct(chat_id, raw_text[5:].strip())
        elif text_lower in ("today summary", "/today_summary"):
            self._cmd_today_summary(chat_id)
        elif text_lower in ("today absent", "/today_absent"):
            self._cmd_today_absent(chat_id)
        elif text_lower in ("dept summary", "/dept_summary"):
            self._cmd_dept_summary(chat_id)
        elif text_lower in ("cache refresh", "/cache_refresh"):
            self._cmd_cache_refresh(chat_id)
        elif text_lower in ("unknown users", "/unknown_users"):
            self._cmd_unknown_users(chat_id)
        elif text_lower in ("pending punches", "/pending_punches"):
            self._cmd_pending_punches(chat_id)
        elif text_lower in ("holiday check", "/holiday_check"):
            self._cmd_holiday_check(chat_id)
        elif text_lower in ("db stats", "/db_stats"):
            self._cmd_db_stats(chat_id)
        elif text_lower in ("late today", "/late_today"):
            self._cmd_late_today(chat_id)
        elif text_lower in ("early exits", "/early_exits"):
            self._cmd_early_exits(chat_id)
        elif text_lower in ("week summary", "/week_summary"):
            self._cmd_week_summary(chat_id)
        elif text_lower in ("month summary", "/month_summary"):
            self._cmd_month_summary(chat_id)
        elif text_lower in ("top absent", "/top_absent"):
            self._cmd_top_absent(chat_id)
        elif text_lower in ("who is in", "/who_is_in"):
            self._cmd_who_is_in(chat_id)
        elif text_lower in ("punch feed", "/punch_feed"):
            self._cmd_punch_feed(chat_id)
        elif text_lower in ("help", "/help", "/start"):
            self._cmd_help(chat_id)

    # ------------------------------------------------------------------ #
    #  Callback query handler (inline keyboard)                            #
    # ------------------------------------------------------------------ #

    def _handle_callback(self, cq: dict):
        chat_id    = str(cq["message"]["chat"]["id"])
        data       = cq.get("data", "")
        message_id = cq["message"]["message_id"]

        # No-op (empty calendar cells, header labels)
        if data == "cal_noop":
            self._answer_callback(cq["id"])
            return

        # Cancel attendance flow
        if data == "att_cancel":
            self._answer_callback(cq["id"], "Cancelled")
            self._state.pop(chat_id, None)
            self._edit_message_text(chat_id, message_id, "❌ Attendance query cancelled.")
            return

        # Date-choice buttons (Today / Yesterday / Custom / Range)
        if data.startswith("att_choice:"):
            self._answer_callback(cq["id"])
            self._handle_att_choice_callback(chat_id, message_id, data)
            return

        # Calendar month navigation (◀/▶)
        if data.startswith("cal_nav:"):
            self._answer_callback(cq["id"])
            self._handle_cal_nav_callback(chat_id, message_id, data)
            return

        # Calendar day selection
        if data.startswith("cal_day:"):
            self._answer_callback(cq["id"])
            self._handle_cal_day_callback(chat_id, message_id, data)
            return

        # Reboot button
        if not data.startswith("reboot:"):
            self._answer_callback(cq["id"])
            return

        target = data[len("reboot:"):]
        self._answer_callback(cq["id"], "Sending reboot command…")

        if target == "ALL":
            ips = self.get_device_ips_fn() if self.get_device_ips_fn else []
            self._edit_message_text(
                chat_id, message_id,
                "🔄 Rebooting <b>all {n} devices</b>…".format(n=len(ips))
            )
            lines = ["🔴 <b>Reboot All Devices</b>", ""]
            for ip in ips:
                try:
                    if self.reboot_device_fn:
                        self.reboot_device_fn(ip)
                    lines.append("✅ <code>{0}</code> — reboot command sent".format(ip))
                except Exception as exc:
                    lines.append("❌ <code>{0}</code>: {1}".format(ip, str(exc)[:60]))
            lines.append("\n⏳ Devices will be offline for ~30 seconds.")
            self._send(chat_id, "\n".join(lines))
        else:
            ip = target
            self._edit_message_text(
                chat_id, message_id,
                "🔄 Rebooting <code>{0}</code>…".format(ip)
            )
            try:
                if self.reboot_device_fn:
                    self.reboot_device_fn(ip)
                self._send(
                    chat_id,
                    "✅ Reboot command sent to <code>{0}</code>.\n"
                    "⏳ Device will be offline for ~30 seconds.".format(ip),
                )
            except Exception as exc:
                self._send(
                    chat_id,
                    "❌ Reboot failed for <code>{0}</code>:\n{1}".format(ip, str(exc)[:120]),
                )

    # ------------------------------------------------------------------ #
    #  Command handlers                                                    #
    # ------------------------------------------------------------------ #

    def _cmd_device_status(self, chat_id: str):
        today_data = {}
        if self.get_today_fn:
            try:
                today_data = self.get_today_fn() or {}
            except Exception:
                pass

        devices = today_data.get("devices", [])

        # Fall back to a live check if cache is empty
        if not devices and self.get_device_status_fn:
            try:
                devices = self.get_device_status_fn() or []
            except Exception:
                pass

        if not devices:
            self._send(chat_id, "⚠️ No device data available.")
            return

        names = {}
        if self.get_device_names_fn:
            try:
                names = self.get_device_names_fn() or {}
            except Exception:
                pass

        lines = ["📡 <b>Device Status</b>", ""]
        for d in devices:
            ip = d.get("ip", "?")
            name = d.get("name") or names.get(ip, "")
            online = d.get("online", False)
            punches = d.get("punches_today", 0)
            users = d.get("user_count", "?")
            icon = "🟢" if online else "🔴"
            header = "{icon} <code>{ip}</code>".format(icon=icon, ip=ip)
            if name:
                header += " — {0}".format(name)
            lines.append(header)
            if online:
                lines.append(
                    "   👤 Users: {u}  |  👆 Punches today: {p}".format(u=users, p=punches)
                )
            else:
                lines.append("   <i>Offline</i>")

        online_count = sum(1 for d in devices if d.get("online", False))
        total_punches = sum(d.get("punches_today", 0) for d in devices)
        lines.append("")
        lines.append(
            "📊 {on}/{tot} online  |  Total punches today: {tp}".format(
                on=online_count, tot=len(devices), tp=total_punches
            )
        )
        self._send(chat_id, "\n".join(lines))

    def _cmd_device_sync(self, chat_id: str):
        self._send(chat_id, "⏳ Starting device sync (time + users)…")

        clock_results = []
        user_results = []
        errors = []

        if self.sync_clocks_fn:
            try:
                clock_results = self.sync_clocks_fn() or []
            except Exception as exc:
                errors.append("Clock sync error: " + str(exc)[:80])

        if self.sync_users_fn:
            try:
                user_results = self.sync_users_fn() or []
            except Exception as exc:
                errors.append("User sync error: " + str(exc)[:80])

        lines = ["🔄 <b>Device Sync Complete</b>", "", "<b>⏰ Clock Sync:</b>"]
        if clock_results:
            for r in clock_results:
                if r.get("ok"):
                    lines.append(
                        "  ✅ <code>{ip}</code> → {t}".format(ip=r["ip"], t=r.get("synced_to", "?"))
                    )
                else:
                    lines.append(
                        "  ❌ <code>{ip}</code>: {e}".format(ip=r["ip"], e=r.get("error", "Failed")[:60])
                    )
        else:
            lines.append("  ⚠️ No clock results")

        if user_results:
            lines.append("")
            lines.append("<b>👥 User Sync:</b>")
            for r in user_results:
                ip = r.get("ip", "?")
                if r.get("ok"):
                    added = r.get("added", 0)
                    note = r.get("note", "")
                    lines.append(
                        "  ✅ <code>{ip}</code>: {added} users pushed{note}".format(
                            ip=ip, added=added,
                            note=" ({0})".format(note) if note else ""
                        )
                    )
                else:
                    lines.append(
                        "  ❌ <code>{ip}</code>: {e}".format(ip=ip, e=r.get("error", "Failed")[:60])
                    )

        for err in errors:
            lines.append("⚠️ " + err)

        self._send(chat_id, "\n".join(lines))

    def _cmd_device_reboot_ask(self, chat_id: str):
        ips = []
        if self.get_device_ips_fn:
            try:
                ips = self.get_device_ips_fn() or []
            except Exception:
                pass

        if not ips:
            self._send(chat_id, "⚠️ No devices configured.")
            return

        names = {}
        if self.get_device_names_fn:
            try:
                names = self.get_device_names_fn() or {}
            except Exception:
                pass

        keyboard = []
        for ip in ips:
            name = names.get(ip, "")
            label = "{0} ({1})".format(name, ip) if name else ip
            keyboard.append([{"text": label, "callback_data": "reboot:{0}".format(ip)}])
        keyboard.append([{"text": "⚡ All Devices", "callback_data": "reboot:ALL"}])

        self._send(
            chat_id,
            "🔴 <b>Device Reboot</b>\nSelect a device to reboot:",
            reply_markup={"inline_keyboard": keyboard},
        )

    def _cmd_device_search_ask(self, chat_id: str):
        self._state[chat_id] = {"state": "awaiting_search_query"}
        self._send(chat_id, "🔍 <b>Employee Search</b>\nEnter employee name or badge number:")

    def _handle_search_query(self, chat_id: str, query: str):
        if not query:
            self._send(chat_id, "⚠️ Empty query — please try again with <code>user search</code>.")
            return
        if self.search_employee_fn:
            try:
                result = self.search_employee_fn(query)
                self._send(chat_id, result)
            except Exception as exc:
                self._send(chat_id, "❌ Search error: {0}".format(str(exc)[:100]))
        else:
            self._send(chat_id, "⚠️ Search is not available right now.")

    def _cmd_user_direct(self, chat_id: str, query: str):
        """Handle direct 'user <name_or_badge>' lookups, e.g. 'user 1024' or 'user showkath'."""
        if not query:
            self._send(chat_id, "⚠️ Please provide a name or badge number, e.g. <code>user 1024</code>.")
            return
        if self.get_employee_punches_fn:
            try:
                result = self.get_employee_punches_fn(query)
                self._send(chat_id, result)
            except Exception as exc:
                self._send(chat_id, "❌ Lookup error: {0}".format(str(exc)[:100]))
        else:
            self._send(chat_id, "⚠️ Employee lookup is not available right now.")

    def _cmd_help(self, chat_id: str):
        text = (
            "📋 <b>Available Commands</b>\n\n"
            "<b>📊 Attendance</b>\n"
            "• <code>today summary</code> — Present/absent/total counts\n"
            "• <code>today absent</code> — Full absent list by department\n"
            "• <code>dept summary</code> — Per-department breakdown\n"
            "• <code>week summary</code> — Present/absent counts Mon–today\n"
            "• <code>month summary</code> — Department attendance % this month\n"
            "• <code>cache refresh</code> — Trigger an immediate data refresh\n\n"
            "<b>👤 Employees</b>\n"
            "• <code>user attendance</code> — Date-wise punches (calendar picker)\n"
            "• <code>user search</code> — Search by name/badge with punch timings today\n"
            "• <code>user &lt;name/badge&gt;</code> — Direct lookup, e.g. <code>user 1024</code>\n"
            "• <code>user report</code> — Punch times for an employee today\n"
            "• <code>late today</code> — Employees who arrived late\n"
            "• <code>early exits</code> — Employees who left early\n"
            "• <code>who is in</code> — Employees currently in the office\n"
            "• <code>top absent</code> — Top 10 most absent this month\n\n"
            "<b>📡 Devices</b>\n"
            "• <code>device status</code> — Show all device statuses\n"
            "• <code>device sync</code> — Sync time &amp; users across all devices\n"
            "• <code>device reboot</code> — Reboot a device (choose from list)\n"
            "• <code>unknown users</code> — Badge IDs not in employee list\n"
            "• <code>punch feed</code> — 20 most recent punches\n\n"
            "<b>🗄️ System</b>\n"
            "• <code>pending punches</code> — Punch corrections awaiting approval\n"
            "• <code>holiday check</code> — Today's holiday &amp; next 30 days\n"
            "• <code>db stats</code> — Employee count, records, DB size\n"
        )
        self._send(chat_id, text)

    # ------------------------------------------------------------------ #
    #  New command handlers                                                #
    # ------------------------------------------------------------------ #

    def _cmd_today_summary(self, chat_id: str):
        today_data = {}
        if self.get_today_fn:
            try:
                today_data = self.get_today_fn() or {}
            except Exception:
                pass
        if not today_data:
            self._send(chat_id, "⚠️ No attendance data available yet. Try <code>cache refresh</code>.")
            return
        date_str      = today_data.get("date", "today")
        present       = today_data.get("present_count", 0)
        absent        = today_data.get("absent_count", 0)
        total         = today_data.get("working_today", present + absent)
        punch_count   = today_data.get("punch_count", 0)
        cache_age     = today_data.get("cache_age_secs")
        refreshing    = today_data.get("refreshing", False)

        age_txt = ""
        if cache_age is not None:
            mins, secs = divmod(int(cache_age), 60)
            age_txt = " (data {0}m {1}s old)".format(mins, secs) if mins else " (data {0}s old)".format(secs)
        if refreshing:
            age_txt += " 🔄"

        lines = [
            "📊 <b>Today's Summary — {0}</b>".format(date_str),
            "",
            "✅ Present:  <b>{0}</b>".format(present),
            "❌ Absent:   <b>{0}</b>".format(absent),
            "👥 Total:    <b>{0}</b>".format(total),
            "👆 Punches:  <b>{0}</b>".format(punch_count),
        ]
        if age_txt:
            lines.append("")
            lines.append("<i>{0}</i>".format(age_txt.strip()))
        self._send(chat_id, "\n".join(lines))

    def _cmd_today_absent(self, chat_id: str):
        today_data = {}
        if self.get_today_fn:
            try:
                today_data = self.get_today_fn() or {}
            except Exception:
                pass
        absent = today_data.get("absent", [])
        if not today_data:
            self._send(chat_id, "⚠️ No attendance data available yet.")
            return
        if not absent:
            self._send(chat_id, "✅ <b>No absences today!</b>")
            return

        dept_groups: dict = {}
        for emp in absent:
            dept_groups.setdefault(emp.get("dept", "Other"), []).append(emp.get("name", "?"))

        lines = [
            "❌ <b>Absent Today — {0}</b>  ({1} employees)".format(
                today_data.get("date", "today"), len(absent)),
            "",
        ]
        for dept in sorted(dept_groups):
            names = sorted(dept_groups[dept])
            lines.append("<b>{0}</b> ({1})".format(dept, len(names)))
            for name in names:
                lines.append("  · {0}".format(name))
            lines.append("")
        self._send(chat_id, "\n".join(lines))

    def _cmd_dept_summary(self, chat_id: str):
        today_data = {}
        if self.get_today_fn:
            try:
                today_data = self.get_today_fn() or {}
            except Exception:
                pass
        if not today_data:
            self._send(chat_id, "⚠️ No attendance data available yet.")
            return

        present = today_data.get("present", [])
        absent  = today_data.get("absent",  [])

        dept_present: dict = {}
        dept_absent:  dict = {}
        for emp in present:
            dept = emp.get("dept", "Other")
            dept_present[dept] = dept_present.get(dept, 0) + 1
        for emp in absent:
            dept = emp.get("dept", "Other")
            dept_absent[dept] = dept_absent.get(dept, 0) + 1

        all_depts = sorted(set(list(dept_present) + list(dept_absent)))
        if not all_depts:
            self._send(chat_id, "⚠️ No department data found.")
            return

        lines = [
            "🏢 <b>Department Summary — {0}</b>".format(today_data.get("date", "today")),
            "",
        ]
        for dept in all_depts:
            p = dept_present.get(dept, 0)
            a = dept_absent.get(dept, 0)
            t = p + a
            pct = int(round(100 * p / t)) if t else 0
            bar = "█" * (pct // 10) + "░" * (10 - pct // 10)
            lines.append(
                "<b>{dept}</b>  ✅{p} ❌{a} / {t}  [{bar}] {pct}%".format(
                    dept=dept, p=p, a=a, t=t, bar=bar, pct=pct)
            )
        self._send(chat_id, "\n".join(lines))

    def _cmd_cache_refresh(self, chat_id: str):
        ok = False
        if self.cache_refresh_fn:
            try:
                ok = self.cache_refresh_fn()
            except Exception:
                pass
        if ok:
            self._send(chat_id, "🔄 Cache refresh started. Data will update in ~30 seconds.")
        else:
            self._send(chat_id, "❌ Could not start cache refresh.")

    def _cmd_user_report_ask(self, chat_id: str):
        self._state[chat_id] = {"state": "awaiting_user_report_query"}
        self._send(chat_id, "👤 <b>User Report</b>\nEnter employee name or badge number:")

    def _handle_user_report_query(self, chat_id: str, query: str):
        if not query:
            self._send(chat_id, "⚠️ Empty input — please try again with <code>user report</code>.")
            return
        if self.get_employee_punches_fn:
            try:
                result = self.get_employee_punches_fn(query)
                self._send(chat_id, result)
            except Exception as exc:
                self._send(chat_id, "❌ Report error: {0}".format(str(exc)[:100]))
        else:
            self._send(chat_id, "⚠️ User report is not available right now.")

    def _cmd_unknown_users(self, chat_id: str):
        users = []
        if self.get_unknown_users_fn:
            try:
                users = self.get_unknown_users_fn() or []
            except Exception:
                pass
        if not users:
            self._send(chat_id, "✅ No unresolved unknown users on any device.")
            return
        lines = [
            "⚠️ <b>Unknown Users</b>  ({0} unresolved)".format(len(users)),
            "",
        ]
        for u in users[:self._MAX_LIST_ITEMS]:   # cap to stay within Telegram message limit
            lines.append(
                "• UID <code>{uid}</code> — <code>{ip}</code>  <i>{seen}</i>".format(
                    uid=u.get("uid", "?"),
                    ip=u.get("device_ip", "?"),
                    seen=(u.get("seen_at", "") or "")[:16],
                )
            )
        if len(users) > self._MAX_LIST_ITEMS:
            lines.append("\n… and {0} more.".format(len(users) - self._MAX_LIST_ITEMS))
        self._send(chat_id, "\n".join(lines))

    def _cmd_pending_punches(self, chat_id: str):
        pending = []
        if self.get_pending_punches_fn:
            try:
                pending = self.get_pending_punches_fn() or []
            except Exception:
                pass
        if not pending:
            self._send(chat_id, "✅ No punch-correction requests are pending.")
            return
        lines = [
            "🎫 <b>Pending Punch Requests</b>  ({0})".format(len(pending)),
            "",
        ]
        for req in pending[:self._MAX_LIST_ITEMS]:
            badge = req.get("badge", "?")
            name  = req.get("employee_name") or req.get("name", "")
            ts    = (req.get("punch_time") or "")[:16]
            lines.append(
                "• <b>{name}</b> ({badge})  🕐 {ts}".format(
                    name=name or badge, badge=badge, ts=ts)
            )
        if len(pending) > self._MAX_LIST_ITEMS:
            lines.append("\n… and {0} more.".format(len(pending) - self._MAX_LIST_ITEMS))
        self._send(chat_id, "\n".join(lines))

    def _cmd_holiday_check(self, chat_id: str):
        from datetime import date as _date
        today = _date.today()
        upcoming = []
        if self.get_upcoming_holidays_fn:
            try:
                upcoming = self.get_upcoming_holidays_fn() or []
            except Exception:
                pass

        # Check if today is a holiday
        today_str = today.strftime("%Y-%m-%d")
        today_holidays = [h for h in upcoming if h.get("date", "") <= today_str <= h.get("date_end", today_str)]
        future_holidays = [h for h in upcoming if h.get("date", "") > today_str]

        lines = ["📅 <b>Holiday Check — {0}</b>".format(today.strftime("%d %b %Y")), ""]
        if today_holidays:
            for h in today_holidays:
                lines.append("🎉 <b>Today is a holiday!</b>  {0}".format(h.get("label", "")))
        else:
            lines.append("🗓️ Today is a regular working day.")

        if future_holidays:
            lines.append("")
            lines.append("<b>Upcoming holidays (next 30 days):</b>")
            for h in future_holidays[:10]:
                d_start = h.get("date", "")
                d_end   = h.get("date_end", d_start)
                span    = " – {0}".format(d_end) if d_end and d_end != d_start else ""
                lines.append("  📌 {start}{span}  {label}".format(
                    start=d_start, span=span, label=h.get("label", "")))
        else:
            lines.append("")
            lines.append("No upcoming holidays in the next 30 days.")
        self._send(chat_id, "\n".join(lines))

    def _cmd_db_stats(self, chat_id: str):
        stats = {}
        if self.get_db_stats_fn:
            try:
                stats = self.get_db_stats_fn() or {}
            except Exception:
                pass
        if not stats or "error" in stats:
            self._send(chat_id, "❌ Could not retrieve database statistics.")
            return
        lines = [
            "🗄️ <b>Database Statistics</b>",
            "",
            "👥 Employees:     <b>{0}</b> total  ({1} active)".format(
                stats.get("employees", "?"), stats.get("active", "?")),
            "👆 Punch records: <b>{0:,}</b>".format(stats.get("punch_records", 0)),
            "❓ Unknown users: <b>{0}</b>".format(stats.get("unknown_users", 0)),
            "💾 DB size:       <b>{0} MB</b>".format(stats.get("size_mb", "?")),
            "📅 First punch:   {0}".format((stats.get("first_punch") or "—")[:10]),
            "📅 Last punch:    {0}".format((stats.get("last_punch") or "—")[:10]),
        ]
        self._send(chat_id, "\n".join(lines))

    # ------------------------------------------------------------------ #
    #  User Attendance — calendar-based date picker                        #
    # ------------------------------------------------------------------ #

    def _cmd_user_attendance_ask(self, chat_id: str):
        self._state[chat_id] = {"state": "awaiting_attendance_query"}
        self._send(
            chat_id,
            "📅 <b>User Attendance</b>\nEnter employee name or badge number:",
        )

    def _handle_attendance_query(self, chat_id: str, query: str):
        if not query:
            self._send(chat_id, "⚠️ Empty query — try again with <code>user attendance</code>.")
            return
        emp = None
        if self.find_employee_fn:
            try:
                emp = self.find_employee_fn(query)
            except Exception as exc:
                self._send(chat_id, "❌ Lookup error: {0}".format(str(exc)[:100]))
                return
        if not emp:
            self._send(chat_id, "🔍 No employee found matching <b>{0}</b>".format(query))
            return
        badge = emp.get("badge", "")
        name  = emp.get("name",  "")
        dept  = emp.get("dept",  "")
        self._state[chat_id] = {"state": "awaiting_date_choice", "badge": badge, "name": name, "dept": dept}
        keyboard = {
            "inline_keyboard": [
                [
                    {"text": "📅 Today",       "callback_data": "att_choice:{b}:T".format(b=badge)},
                    {"text": "📅 Yesterday",   "callback_data": "att_choice:{b}:Y".format(b=badge)},
                ],
                [
                    {"text": "📆 Custom Date", "callback_data": "att_choice:{b}:C".format(b=badge)},
                    {"text": "📆 Date Range",  "callback_data": "att_choice:{b}:R".format(b=badge)},
                ],
                [{"text": "❌ Cancel", "callback_data": "att_cancel"}],
            ]
        }
        self._send(
            chat_id,
            "👤 <b>{name}</b> ({badge}) — {dept}\nChoose a date option:".format(
                name=name, badge=badge, dept=dept),
            reply_markup=keyboard,
        )

    def _handle_att_choice_callback(self, chat_id: str, message_id: int, data: str):
        # data format: "att_choice:{badge}:{choice}" where choice = T/Y/C/R
        parts  = data.split(":", 2)
        if len(parts) < 3:
            return
        badge  = parts[1]
        choice = parts[2]

        # Resolve name/dept from state (set when employee was found)
        st    = self._state.get(chat_id, {})
        name  = st.get("name", badge)
        dept  = st.get("dept", "")

        from datetime import date as _date, timedelta as _td
        today     = _date.today()
        yesterday = today - _td(days=1)

        if choice == "T":
            self._state.pop(chat_id, None)
            self._edit_message_text(chat_id, message_id,
                "⏳ Fetching attendance for {name} — {d}…".format(
                    name=name, d=today.strftime("%d %b %Y")))
            result = self._fetch_single_day(badge, name, dept, today.strftime("%Y-%m-%d"))
            self._edit_message_text(chat_id, message_id, result)

        elif choice == "Y":
            self._state.pop(chat_id, None)
            self._edit_message_text(chat_id, message_id,
                "⏳ Fetching attendance for {name} — {d}…".format(
                    name=name, d=yesterday.strftime("%d %b %Y")))
            result = self._fetch_single_day(badge, name, dept, yesterday.strftime("%Y-%m-%d"))
            self._edit_message_text(chat_id, message_id, result)

        elif choice == "C":
            # Show calendar — single mode
            year, month = today.year, today.month
            self._state[chat_id] = {"state": "cal_single", "badge": badge, "name": name, "dept": dept}
            kb   = self._make_calendar_keyboard(year, month, "S", badge)
            text = "📅 <b>Select date</b>\n👤 {name} ({badge})".format(name=name, badge=badge)
            self._edit_message_text_and_keyboard(chat_id, message_id, text,
                                                 reply_markup={"inline_keyboard": kb})

        elif choice == "R":
            # Show calendar — range-from mode
            year, month = today.year, today.month
            self._state[chat_id] = {"state": "cal_range_from", "badge": badge, "name": name, "dept": dept}
            kb   = self._make_calendar_keyboard(year, month, "F", badge)
            text = "📅 <b>Select start date</b>\n👤 {name} ({badge})".format(name=name, badge=badge)
            self._edit_message_text_and_keyboard(chat_id, message_id, text,
                                                 reply_markup={"inline_keyboard": kb})

    def _handle_cal_nav_callback(self, chat_id: str, message_id: int, data: str):
        # data format: "cal_nav:{year}:{month}:{mode}:{badge}:{range_from}"
        parts = data.split(":")
        # parts: ['cal_nav', year, month, mode, badge, range_from(may be empty)]
        if len(parts) < 6:
            return
        try:
            year  = int(parts[1])
            month = int(parts[2])
        except ValueError:
            return
        mode       = parts[3]
        badge      = parts[4]
        range_from = parts[5] if len(parts) > 5 else ""

        st   = self._state.get(chat_id, {})
        name = st.get("name", badge)

        if mode == "S":
            title = "📅 <b>Select date</b>\n👤 {name} ({badge})".format(name=name, badge=badge)
        elif mode == "F":
            title = "📅 <b>Select start date</b>\n👤 {name} ({badge})".format(name=name, badge=badge)
        else:  # "T"
            title = "📅 <b>Select end date</b>\n👤 {name} ({badge})\n⏩ From: {rf}".format(
                name=name, badge=badge, rf=range_from)

        kb = self._make_calendar_keyboard(year, month, mode, badge, range_from)
        self._edit_message_text_and_keyboard(chat_id, message_id, title,
                                             reply_markup={"inline_keyboard": kb})

    def _handle_cal_day_callback(self, chat_id: str, message_id: int, data: str):
        # data format: "cal_day:{date}:{mode}:{badge}:{range_from}"
        # date = YYYY-MM-DD
        parts = data.split(":")
        # parts: ['cal_day', 'YYYY-MM-DD', mode, badge, range_from]
        if len(parts) < 5:
            return
        selected_date = parts[1]
        mode          = parts[2]
        badge         = parts[3]
        range_from    = parts[4] if len(parts) > 4 else ""

        st   = self._state.get(chat_id, {})
        name = st.get("name", badge)
        dept = st.get("dept", "")

        from datetime import date as _date, timedelta as _td
        today = _date.today()

        if mode == "S":
            # Single day — fetch and show
            self._state.pop(chat_id, None)
            self._edit_message_text(chat_id, message_id,
                "⏳ Fetching attendance for {name} — {d}…".format(name=name, d=selected_date))
            result = self._fetch_single_day(badge, name, dept, selected_date)
            self._edit_message_text(chat_id, message_id, result)

        elif mode == "F":
            # Range — got start date; show calendar for end date
            year  = int(selected_date[:4])
            month = int(selected_date[5:7])
            self._state[chat_id] = {
                "state": "cal_range_to", "badge": badge, "name": name, "dept": dept,
                "range_from": selected_date,
            }
            kb   = self._make_calendar_keyboard(year, month, "T", badge, selected_date)
            text = "📅 <b>Select end date</b>\n👤 {name} ({badge})\n⏩ From: {rf}".format(
                name=name, badge=badge, rf=selected_date)
            self._edit_message_text_and_keyboard(chat_id, message_id, text,
                                                 reply_markup={"inline_keyboard": kb})

        elif mode == "T":
            # Range — got end date; fetch range
            self._state.pop(chat_id, None)
            d_from_str = range_from
            d_to_str   = selected_date
            # Ensure from <= to using datetime comparison (robust regardless of format)
            try:
                d_from_dt = datetime.strptime(d_from_str, "%Y-%m-%d")
                d_to_dt   = datetime.strptime(d_to_str,   "%Y-%m-%d")
                if d_from_dt > d_to_dt:
                    d_from_str, d_to_str = d_to_str, d_from_str
            except ValueError:
                pass
            self._edit_message_text(chat_id, message_id,
                "⏳ Fetching attendance for {name} — {df} to {dt}…".format(
                    name=name, df=d_from_str, dt=d_to_str))
            result = self._fetch_range(badge, name, dept, d_from_str, d_to_str)
            self._edit_message_text(chat_id, message_id, result)

    # ------------------------------------------------------------------ #
    #  Calendar keyboard builder                                           #
    # ------------------------------------------------------------------ #

    @staticmethod
    def _make_calendar_keyboard(year: int, month: int, mode: str,
                                badge: str, range_from: str = "") -> list:
        """
        Build a month-grid inline keyboard.
        mode: 'S'=single, 'F'=range-from, 'T'=range-to
        Each day button carries callback_data "cal_day:YYYY-MM-DD:{mode}:{badge}:{range_from}".
        Navigation buttons carry "cal_nav:{year}:{month}:{mode}:{badge}:{range_from}".
        Empty cells and header labels carry "cal_noop".
        """
        import calendar as _cal

        prev_y, prev_m = (year, month - 1) if month > 1 else (year - 1, 12)
        next_y, next_m = (year, month + 1) if month < 12 else (year + 1, 1)
        rf = range_from  # short alias

        def nav_data(y, m):
            return "cal_nav:{y}:{m}:{mode}:{badge}:{rf}".format(
                y=y, m=m, mode=mode, badge=badge, rf=rf)

        month_label = datetime(year, month, 1).strftime("%b %Y")

        keyboard = [
            [
                {"text": "◀", "callback_data": nav_data(prev_y, prev_m)},
                {"text": month_label, "callback_data": "cal_noop"},
                {"text": "▶", "callback_data": nav_data(next_y, next_m)},
            ],
            [{"text": d, "callback_data": "cal_noop"}
             for d in ["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]],
        ]

        # first_day.weekday(): 0=Mon … 6=Sun → maps directly to column index
        first_col   = datetime(year, month, 1).weekday()
        total_days  = _cal.monthrange(year, month)[1]
        week        = [{"text": " ", "callback_data": "cal_noop"}] * first_col

        for day in range(1, total_days + 1):
            date_str = "{y:04d}-{m:02d}-{d:02d}".format(y=year, m=month, d=day)
            week.append({
                "text": str(day),
                "callback_data": "cal_day:{date}:{mode}:{badge}:{rf}".format(
                    date=date_str, mode=mode, badge=badge, rf=rf),
            })
            if len(week) == 7:
                keyboard.append(week)
                week = []

        if week:
            week.extend([{"text": " ", "callback_data": "cal_noop"}] * (7 - len(week)))
            keyboard.append(week)

        keyboard.append([{"text": "❌ Cancel", "callback_data": "att_cancel"}])
        return keyboard

    # ------------------------------------------------------------------ #
    #  Attendance result formatters                                        #
    # ------------------------------------------------------------------ #

    @staticmethod
    def _fmt_time(ts: str) -> str:
        try:
            return datetime.strptime(ts, "%Y-%m-%d %H:%M:%S").strftime("%I:%M %p")
        except Exception:
            return ts[:16]

    def _fetch_single_day(self, badge: str, name: str, dept: str, date_str: str) -> str:
        records = []
        if self.get_punch_records_fn:
            try:
                records = self.get_punch_records_fn(badge, date_str, date_str) or []
            except Exception as exc:
                return "❌ Error fetching punches: {0}".format(str(exc)[:100])
        try:
            d = datetime.strptime(date_str, "%Y-%m-%d")
            day_label = d.strftime("%A %d %b %Y")
        except Exception:
            day_label = date_str
        lines = [
            "👤 <b>{name}</b> ({badge}) — {dept}".format(name=name, badge=badge, dept=dept),
            "📅 {0}".format(day_label),
            "",
        ]
        if not records:
            lines.append("❌ No punches recorded.")
        else:
            lines.append("✅ <b>{n} punch(es):</b>".format(n=len(records)))
            for r in records:
                lines.append("  🕐 {0}  (<code>{1}</code>)".format(
                    self._fmt_time(r.get("punch_time", "")),
                    r.get("device_ip", "?")))
            lines.append("")
            if len(records) >= 2:
                lines.append("⏩ First: {0}   Last: {1}".format(
                    self._fmt_time(records[0]["punch_time"]),
                    self._fmt_time(records[-1]["punch_time"])))
        return "\n".join(lines)

    def _fetch_range(self, badge: str, name: str, dept: str,
                     d_from_str: str, d_to_str: str) -> str:
        records = []
        if self.get_punch_records_fn:
            try:
                records = self.get_punch_records_fn(badge, d_from_str, d_to_str) or []
            except Exception as exc:
                return "❌ Error fetching punches: {0}".format(str(exc)[:100])

        from datetime import date as _date, timedelta as _td
        d_from     = datetime.strptime(d_from_str, "%Y-%m-%d").date()
        d_to       = datetime.strptime(d_to_str,   "%Y-%m-%d").date()
        total_days = (d_to - d_from).days + 1

        # Group records by date
        by_date: Dict[str, list] = {}
        for r in records:
            key = r["punch_time"][:10]
            by_date.setdefault(key, []).append(r)

        header = [
            "👤 <b>{name}</b> ({badge}) — {dept}".format(name=name, badge=badge, dept=dept),
            "📅 {df} – {dt}  ({n} days)".format(
                df=d_from.strftime("%d %b"),
                dt=d_to.strftime("%d %b %Y"),
                n=total_days),
            "",
        ]

        present = 0
        absent  = 0
        weekend = 0
        lines   = []
        d       = d_from
        compact = total_days > 14

        while d <= d_to:
            dow       = d.weekday()       # 0=Mon…6=Sun; Fri=4, Sat=5
            day_key   = d.strftime("%Y-%m-%d")
            day_recs  = by_date.get(day_key, [])
            is_weekend = dow in (4, 5)    # Fri/Sat = weekend in Middle East

            if is_weekend:
                weekend += 1
            else:
                if day_recs:
                    present += 1
                else:
                    absent += 1

            if compact:
                if day_recs:
                    t_in  = self._fmt_time(day_recs[0]["punch_time"])
                    t_out = self._fmt_time(day_recs[-1]["punch_time"]) if len(day_recs) > 1 else "—     "
                    lines.append("{ds}  {t_in}  {t_out}  ({n})".format(
                        ds=d.strftime("%d %b %a"), t_in=t_in, t_out=t_out, n=len(day_recs)))
                elif not is_weekend:
                    lines.append("{ds}  —         —         0".format(ds=d.strftime("%d %b %a")))
            else:
                day_str = d.strftime("%a %d %b")
                if is_weekend:
                    lines.append("📆 {day}  🏖 Weekend".format(day=day_str))
                elif not day_recs:
                    lines.append("📆 {day}  ❌  No punches".format(day=day_str))
                elif len(day_recs) == 1:
                    lines.append("📆 {day}  ✅  IN {t}  (1 punch — no out)".format(
                        day=day_str, t=self._fmt_time(day_recs[0]["punch_time"])))
                else:
                    lines.append("📆 {day}  ✅  IN {t_in}  OUT {t_out}  ({n} punches)".format(
                        day=day_str,
                        t_in=self._fmt_time(day_recs[0]["punch_time"]),
                        t_out=self._fmt_time(day_recs[-1]["punch_time"]),
                        n=len(day_recs)))
            d += _td(days=1)

        summary = "\nSummary: {p} present, {a} absent".format(p=present, a=absent)
        if weekend:
            summary += ", {w} weekend".format(w=weekend)

        return "\n".join(header + lines + [summary])

    # ------------------------------------------------------------------ #
    #  New attendance-insight command handlers                             #
    # ------------------------------------------------------------------ #

    def _cmd_late_today(self, chat_id: str):
        late = []
        if self.get_late_today_fn:
            try:
                late = self.get_late_today_fn() or []
            except Exception as exc:
                self._send(chat_id, "❌ Error: {0}".format(str(exc)[:100]))
                return
        if not late:
            self._send(chat_id, "✅ <b>No late arrivals today!</b>")
            return
        from datetime import date as _date
        lines = [
            "⏰ <b>Late Today — {0}</b>  ({1} employees)".format(
                _date.today().strftime("%d %b %Y"), len(late)),
            "",
        ]
        for emp in late[:self._MAX_LIST_ITEMS]:
            lines.append(
                "• <b>{name}</b> ({badge})  🕐 {t}  (+{m}m late)".format(
                    name=emp.get("name", "?"),
                    badge=emp.get("badge", "?"),
                    t=emp.get("first_punch", "?"),
                    m=emp.get("minutes_late", 0),
                )
            )
        if len(late) > self._MAX_LIST_ITEMS:
            lines.append("\n… and {0} more.".format(len(late) - self._MAX_LIST_ITEMS))
        self._send(chat_id, "\n".join(lines))

    def _cmd_early_exits(self, chat_id: str):
        exits = []
        if self.get_early_exits_fn:
            try:
                exits = self.get_early_exits_fn() or []
            except Exception as exc:
                self._send(chat_id, "❌ Error: {0}".format(str(exc)[:100]))
                return
        if not exits:
            self._send(chat_id, "✅ <b>No early exits today!</b> (or shift hasn't ended yet)")
            return
        from datetime import date as _date
        lines = [
            "🏃 <b>Early Exits — {0}</b>  ({1} employees)".format(
                _date.today().strftime("%d %b %Y"), len(exits)),
            "",
        ]
        for emp in exits[:self._MAX_LIST_ITEMS]:
            lines.append(
                "• <b>{name}</b> ({badge})  🕐 left {t}  (-{m}m early)".format(
                    name=emp.get("name", "?"),
                    badge=emp.get("badge", "?"),
                    t=emp.get("last_punch", "?"),
                    m=emp.get("minutes_early", 0),
                )
            )
        if len(exits) > self._MAX_LIST_ITEMS:
            lines.append("\n… and {0} more.".format(len(exits) - self._MAX_LIST_ITEMS))
        self._send(chat_id, "\n".join(lines))

    def _cmd_week_summary(self, chat_id: str):
        days = []
        if self.get_week_summary_fn:
            try:
                days = self.get_week_summary_fn() or []
            except Exception as exc:
                self._send(chat_id, "❌ Error: {0}".format(str(exc)[:100]))
                return
        if not days:
            self._send(chat_id, "⚠️ No data available for this week.")
            return
        lines = ["📆 <b>Week Summary</b>", ""]
        for d in days:
            p   = d.get("present", 0)
            a   = d.get("absent",  0)
            t   = d.get("total",   0)
            pct = int(round(100 * p / t)) if t else 0
            bar = "█" * (pct // 10) + "░" * (10 - pct // 10)
            lines.append(
                "<b>{wd} {dt}</b>  ✅{p} ❌{a}  [{bar}] {pct}%".format(
                    wd=d.get("weekday", "?"),
                    dt=(d.get("date", "") or "")[-5:].replace("-", "/"),
                    p=p, a=a, bar=bar, pct=pct)
            )
        self._send(chat_id, "\n".join(lines))

    def _cmd_month_summary(self, chat_id: str):
        depts = []
        if self.get_month_dept_summary_fn:
            try:
                depts = self.get_month_dept_summary_fn() or []
            except Exception as exc:
                self._send(chat_id, "❌ Error: {0}".format(str(exc)[:100]))
                return
        if not depts:
            self._send(chat_id, "⚠️ No data available for this month.")
            return
        from datetime import date as _date
        today = _date.today()
        lines = [
            "🏢 <b>Month Summary — {0}</b>".format(today.strftime("%B %Y")),
            "",
        ]
        for d in depts:
            pct = d.get("pct", 0)
            bar = "█" * (pct // 10) + "░" * (10 - pct // 10)
            lines.append(
                "<b>{dept}</b>  {n} staff  avg {avg}d/{wd}d  [{bar}] {pct}%".format(
                    dept=d.get("dept", "?"),
                    n=d.get("total", 0),
                    avg=d.get("avg_present", 0),
                    wd=d.get("working_days", 0),
                    bar=bar,
                    pct=pct,
                )
            )
        self._send(chat_id, "\n".join(lines))

    def _cmd_top_absent(self, chat_id: str):
        emps = []
        if self.get_top_absent_fn:
            try:
                emps = self.get_top_absent_fn() or []
            except Exception as exc:
                self._send(chat_id, "❌ Error: {0}".format(str(exc)[:100]))
                return
        if not emps:
            self._send(chat_id, "✅ <b>No absences recorded this month!</b>")
            return
        from datetime import date as _date
        lines = [
            "📉 <b>Top Absent — {0}</b>".format(_date.today().strftime("%B %Y")),
            "",
        ]
        for i, emp in enumerate(emps[:10], 1):
            lines.append(
                "{i}. <b>{name}</b> ({badge}) — {dept}\n"
                "   ❌ {a} absent days  ✅ {p} present days".format(
                    i=i,
                    name=emp.get("name", "?"),
                    badge=emp.get("badge", "?"),
                    dept=emp.get("dept", "?"),
                    a=emp.get("days_absent", 0),
                    p=emp.get("days_present", 0),
                )
            )
        self._send(chat_id, "\n".join(lines))

    def _cmd_who_is_in(self, chat_id: str):
        emps = []
        if self.get_who_is_in_fn:
            try:
                emps = self.get_who_is_in_fn() or []
            except Exception as exc:
                self._send(chat_id, "❌ Error: {0}".format(str(exc)[:100]))
                return
        from datetime import date as _date
        count = len(emps)
        if not emps:
            self._send(chat_id, "🏢 <b>Who Is In — {0}</b>\n\nNo one is currently logged in.".format(
                _date.today().strftime("%d %b %Y")))
            return
        lines = [
            "🏢 <b>Who Is In — {0}</b>  ({1} employees)".format(
                _date.today().strftime("%d %b %Y"), count),
            "",
        ]
        for emp in emps[:self._MAX_LIST_ITEMS]:
            last_t = self._fmt_time(emp.get("last_punch", ""))
            lines.append(
                "• <b>{name}</b> ({badge})  🕐 {t}".format(
                    name=emp.get("name", "?"),
                    badge=emp.get("badge", "?"),
                    t=last_t,
                )
            )
        if count > self._MAX_LIST_ITEMS:
            lines.append("\n… and {0} more.".format(count - self._MAX_LIST_ITEMS))
        self._send(chat_id, "\n".join(lines))

    def _cmd_punch_feed(self, chat_id: str):
        feed = []
        if self.get_punch_feed_fn:
            try:
                feed = self.get_punch_feed_fn() or []
            except Exception as exc:
                self._send(chat_id, "❌ Error: {0}".format(str(exc)[:100]))
                return
        if not feed:
            self._send(chat_id, "⚠️ No recent punch records found.")
            return
        lines = ["👆 <b>Recent Punches (last {0})</b>".format(len(feed)), ""]
        for r in feed:
            name_str = r.get("name") or r.get("badge", "?")
            ts       = self._fmt_time(r.get("punch_time", ""))
            ip       = r.get("device_ip", "?")
            lines.append(
                "🕐 {ts}  <b>{name}</b>  <code>{ip}</code>".format(
                    ts=ts, name=name_str, ip=ip)
            )
        self._send(chat_id, "\n".join(lines))




_CATEGORY_MAP = {
    "TEACHING":       "Teachers",
    "ADMIN":          "Admin & Support",
    "SUPPORT":        "Admin & Support",
    "DRIVER":         "Drivers & Conductors",
    "CONDUCTOR":      "Drivers & Conductors",
    "CLEANING STAFF": "Cleaners",
}
_CATEGORY_ORDER = ["Teachers", "Admin & Support", "Drivers & Conductors", "Cleaners"]
_CAT_COLORS = {
    "Teachers":           "1F4E79",
    "Admin & Support":    "375623",
    "Drivers & Conductors": "7B3F00",
    "Cleaners":           "4A235A",
}


def _build_absent_xlsx(absent: list, date_str: str) -> bytes:
    """
    Build an XLSX workbook for the daily absent report.
    Columns: No. | Name | Date | Timetable
    Rows grouped by dept category (Teachers / Admin+Support / etc.)
    Returns raw bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    def _thin():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    wb = Workbook()

    # ---------- single combined sheet ----------
    ws = wb.active
    ws.title = "Absent"
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value="Daily Absent Report — {0}".format(date_str))
    c.font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Sub-title
    ws.cell(row=2, column=1,
            value="Generated: {0}".format(datetime.now().strftime("%d %b %Y %H:%M"))
            ).font = Font(name="Arial", size=9, italic=True, color="888888")

    # Header row
    headers = ["No.", "Name", "Date", "Timetable"]
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=hdr)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", start_color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin()
    ws.row_dimensions[4].height = 18

    # Group by category
    buckets: Dict[str, list] = {}
    for emp in absent:
        dept = (emp.get("dept") or "").upper()
        cat = _CATEGORY_MAP.get(dept, "Others")
        buckets.setdefault(cat, []).append(emp)

    ordered_cats = [c for c in _CATEGORY_ORDER if c in buckets]
    ordered_cats += sorted(k for k in buckets if k not in ordered_cats)

    row = 5
    for cat in ordered_cats:
        emps = sorted(buckets[cat], key=lambda e: e.get("name", ""))
        cat_color = _CAT_COLORS.get(cat, "444444")

        # Category separator row
        ws.merge_cells("A{r}:D{r}".format(r=row))
        c = ws.cell(row=row, column=1, value="{cat}  ({n} absent)".format(cat=cat, n=len(emps)))
        c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=cat_color)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = _thin()
        ws.row_dimensions[row].height = 20
        row += 1

        for i, emp in enumerate(emps):
            bg = "EBF3FB" if i % 2 == 0 else "FFFFFF"
            vals = [
                emp.get("code", ""),
                emp.get("name", ""),
                date_str,
                emp.get("dept", ""),
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name="Arial", size=10)
                c.fill = PatternFill("solid", start_color=bg)
                c.border = _thin()
                c.alignment = Alignment(
                    horizontal="center" if col in (1, 3) else "left",
                    vertical="center",
                )
            row += 1

    # Column widths
    for col, w in zip("ABCD", [12, 34, 14, 22]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()